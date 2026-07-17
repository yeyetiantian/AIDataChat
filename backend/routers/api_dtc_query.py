"""DTC 数据查询 API — 通过 SQL 查询 dtc_info / dtc_trigger 表，支持多 Sheet Excel 导出"""

from __future__ import annotations

import io
import logging
import os
import re
import sqlite3
import sys
from typing import Any

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

logger = logging.getLogger("api_dtc_query")

if getattr(sys, "frozen", False):
    _backend_dir = os.path.dirname(os.path.abspath(sys.executable))
else:
    _backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DB_PATH = os.path.join(_backend_dir, "ai_data.db")

router = APIRouter(prefix="/api/dtc", tags=["dtc"])

# 仅允许查询这些表
_ALLOWED_TABLES = {"dtc_info", "dtc_trigger"}

# SQL 危险关键字（不允许出现）
_DANGEROUS_KW = [
    "DELETE", "DROP", "INSERT", "UPDATE", "ALTER", "CREATE",
    "PRAGMA", "ATTACH", "DETACH", "REINDEX", "REPLACE", "TRUNCATE",
]


class SqlQueryRequest(BaseModel):
    sql: str = Field(..., min_length=7, description="SELECT 查询语句，仅限查询 dtc_info / dtc_trigger 表")


def _get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _validate_sql(sql: str) -> str:
    """校验 SQL：只允许 SELECT，不允许危险操作"""
    stripped = sql.strip()
    if not stripped.upper().startswith("SELECT"):
        raise HTTPException(status_code=400, detail="只允许 SELECT 查询")

    upper_sql = stripped.upper()
    for kw in _DANGEROUS_KW:
        if re.search(rf"\b{re.escape(kw)}\b", upper_sql):
            raise HTTPException(status_code=400, detail=f"查询中包含不允许的关键字: {kw}")

    return stripped


@router.post("/query")
async def query_dtc(req: SqlQueryRequest) -> dict[str, Any]:
    """执行 SQL 查询 dtc_info / dtc_trigger 表，返回列名和数据行"""
    sql = _validate_sql(req.sql)
    conn = _get_conn()
    try:
        cursor = conn.execute(sql)
        rows = [dict(row) for row in cursor.fetchall()]
        columns = [desc[0] for desc in cursor.description]
        logger.info("SQL OK (%d rows): %.200s", len(rows), sql)
        return {
            "success": True,
            "columns": columns,
            "rows": rows,
            "total": len(rows),
        }
    except sqlite3.Error as e:
        logger.warning("SQL ERROR: %s — %.200s", e, sql)
        raise HTTPException(status_code=400, detail=f"SQL 执行错误: {e}")
    finally:
        conn.close()


class SheetData(BaseModel):
    """单个 Sheet 的数据"""
    name: str = Field(..., description="Sheet 名称")
    columns: list[str] = Field(..., description="列名列表")
    rows: list[dict[str, Any]] = Field(..., description="数据行列表")


class DtcExportRequest(BaseModel):
    """DTC 数据导出请求（支持多 Sheet）"""
    sheets: list[SheetData] = Field(..., min_length=1, description="多个 Sheet 的数据")


@router.post("/export")
async def export_dtc_excel(req: DtcExportRequest):
    """将 DTC 查询结果导出为多 Sheet Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl 库，无法生成 Excel")

    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_data in req.sheets:
        ws = wb.create_sheet(title=sheet_data.name[:31])  # Excel Sheet 名称最长 31 字符

        # 写表头
        for col_idx, col_name in enumerate(sheet_data.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写数据
        for row_idx, row_data in enumerate(sheet_data.rows, 2):
            for col_idx, col_name in enumerate(sheet_data.columns, 1):
                value = row_data.get(col_name)
                # 转换类型（openpyxl 不能处理某些类型）
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 自适应列宽（取前 100 行采样）
        for col_idx, col_name in enumerate(sheet_data.columns, 1):
            max_len = len(str(col_name)) * 2  # 中文字符按 2 倍宽度
            sample_rows = sheet_data.rows[:100]
            for row_data in sample_rows:
                val = row_data.get(col_name)
                if val is not None:
                    char_len = len(str(val))
                    # 中文字符按 2 倍估算
                    cn_count = sum(1 for c in str(val) if '一' <= c <= '鿿')
                    max_len = max(max_len, char_len + cn_count)
            # 限制最大列宽 60
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # 写入 BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import datetime
    filename = f"dtc_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )